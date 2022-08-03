import requests
import pandas as pd
from bs4 import BeautifulSoup as BS
from openpyxl import load_workbook

path = "../shuangseqiu.xlsx"
wb = load_workbook(path)
ws = wb['Sheet1']
class TwoColorBall:
    def get_page(self,url):
        headers = {
            "User-Agent":"Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.25 Safari/537.36 Core/1.70.3741.400 QQBrowser/10.5.3863.400",
            "Host":"kaijiang.zhcw.com"
        }
        res =requests.get(url =url,headers=headers)
        text = BS(res.text,"html.parser")
        return text

    def get_text(self):
        '''将结果插入excel'''
        ws.delete_rows(2,ws.max_row)
        for i in range(1,130):
            url = 'http://kaijiang.zhcw.com/zhcw/html/ssq/list_{}.html'.format(i)
            text = self.get_page(url)
            tr_table = text.find_all("tr")
            for tr_num in range(2,len(tr_table)-1):
                td_table = tr_table[tr_num].find_all('td')
                ems = td_table[2].find_all("em")
                ws.cell(ws.max_row + 1, 1, value=td_table[0].get_text())
                ws.cell(ws.max_row, 2,value=td_table[1].get_text())
                ws.cell(ws.max_row, 3, value=ems[0].get_text())
                ws.cell(ws.max_row, 4, value=ems[1].get_text())
                ws.cell(ws.max_row, 5, value=ems[2].get_text())
                ws.cell(ws.max_row, 6, value=ems[3].get_text())
                ws.cell(ws.max_row, 7, value=ems[4].get_text())
                ws.cell(ws.max_row, 8, value=ems[5].get_text())
                ws.cell(ws.max_row, 9, value=ems[6].get_text())
        wb.save("../shuangseqiu.xlsx")
        wb.close()

    def get_count(self,column):
        df = pd.DataFrame(pd.read_excel(path))
        first = df.loc[:,column].value_counts()
        return first.index[0]

    def get_most_frequent_number(self):
        '''
        统计往期每个位置上出现过最多的球
        :return:
        '''
        most_frequent_number = []
        column=['first','second','third','fourth','fifth','sixth','basketball']
        for i in column:
            num = self.get_count(i)
            most_frequent_number.append(num)
        return most_frequent_number

    def get_column_values(self,column):
        '''
        获取column列的所有值
        :param column:
        :return:
        '''
        rows = ws.max_row
        columndata = []
        for i in range(2, rows + 1):
            cellvalue = ws.cell(row=i, column=column).value
            columndata.append(cellvalue)
        columndata.reverse()
        return columndata

    def count_the_number_of_occurrences(self,column):
        '''
        统计红球每个数后面出现的次数
        :return:
        '''
        first = {}
        columndata = self.get_column_values(column)
        for num in range(1,len(columndata)):
            for f_num in range(1,33):
                for a in range(1,33):
                    previous = columndata[num-1] # 当前位数
                    next_num =  columndata[num]  #后一位

                    if previous=='{}'.format('%02d'%f_num) and next_num=='{}'.format('%02d'%a):
                        if previous+next_num in first.keys():
                            first[previous+next_num] += 1
                        else:
                            first[previous+next_num]=1
        return first

    def count_the_number_of_occurrences_basketball(self,column):
        '''
        统计蓝球每个数后面出现的次数
        :return:
        '''
        first = {}
        columndata = self.get_column_values(column)
        for num in range(1,len(columndata)):
            for f_num in range(1,17):
                for a in range(1,17):
                    previous = columndata[num-1] # 当前位数
                    next_num =  columndata[num]  #后一位
                    if previous=='{}'.format('%02d'%f_num) and next_num=='{}'.format('%02d'%a):
                        if previous+next_num in first.keys():
                            first[previous+next_num] += 1
                        else:
                            first[previous+next_num]=1
        return first

    def get_balls_with_the_greatest_chance(self,last_num,num_dict):
        '''
        根据上一期结果返回每个球后出现几率最大的一个球
        :return:
        '''
        dict_num = {}
        for i in range(1,33):
            a = last_num+'%02d'%i
            if a in num_dict.keys():
                dict_num[a]=num_dict[a]
        max_num = max(zip(dict_num.values(), dict_num.keys()))
        return max_num[1]

    def main(self):
        '''计算所有结果'''
        self.get_text()
        b = self.get_most_frequent_number()
        first_dict = self.count_the_number_of_occurrences(3)
        second_dict = self.count_the_number_of_occurrences(4)
        third_dict = self.count_the_number_of_occurrences(5)
        fourth_dict = self.count_the_number_of_occurrences(6)
        fifth_dict = self.count_the_number_of_occurrences(7)
        sixth_dict = self.count_the_number_of_occurrences(8)
        basketball_dict = self.count_the_number_of_occurrences_basketball(9)
        custom_period =2
        first_half =ws.cell(custom_period, 1).value
        last_first = ws.cell(custom_period, 3).value
        last_second = ws.cell(custom_period, 4).value
        last_third = ws.cell(custom_period, 5).value
        last_fourth = ws.cell(custom_period, 6).value
        last_fifth = ws.cell(custom_period,7).value
        last_sixth = ws.cell(custom_period, 8).value
        last_basketball = ws.cell(custom_period, 9).value
        next_first =self.get_balls_with_the_greatest_chance(last_first,first_dict)
        next_second=self.get_balls_with_the_greatest_chance(last_second,second_dict)
        next_third=self.get_balls_with_the_greatest_chance(last_third,third_dict)
        next_fourth=self.get_balls_with_the_greatest_chance(last_fourth,fourth_dict)
        next_fifth=self.get_balls_with_the_greatest_chance(last_fifth,fifth_dict)
        next_sixth=self.get_balls_with_the_greatest_chance(last_sixth,sixth_dict)
        next_basketball=self.get_balls_with_the_greatest_chance(last_basketball,basketball_dict)
        print(f'根据上期{first_half}开奖结果红球1：{next_first[0:2]}后出现最多的是：{next_first[2:4]}')
        print(f'根据上期{first_half}开奖结果红球2：{next_second[0:2]}后出现最多的是：{next_second[2:4]}')
        print(f'根据上期{first_half}开奖结果红球3：{next_third[0:2]}后出现最多的是：{next_third[2:4]}')
        print(f'根据上期{first_half}开奖结果红球4：{next_fourth[0:2]}后出现最多的是：{next_fourth[2:4]}')
        print(f'根据上期{first_half}开奖结果红球5：{next_fifth[0:2]}后出现最多的是：{next_fifth[2:4]}')
        print(f'根据上期{first_half}开奖结果红球6：{next_sixth[0:2]}后出现最多的是：{next_sixth[2:4]}')
        print(f'根据上期{first_half}开奖结果蓝球：{next_basketball[0:2]}后出现最多的是：{next_basketball[2:4]}')
        print('#'*60)
        print(f'每个位置上出现最多的球是：红球：{b[0]},{b[1]},{b[2]},{b[3]},{b[4]},{b[5]}  蓝球：{b[6]}')
        print('#' * 60)
        print(f'##########上期{first_half}开奖结果是：红球：{last_first},{last_second},{last_third},{last_fourth},'
              f'{last_fifth},{last_sixth}  蓝球：{last_basketball}')
        print('#' * 60)
        print(f'根据上期结果预测下期的球：红球：{next_first[2:4]},{next_second[2:4]},{next_third[2:4]},{next_fourth[2:4]},'
              f'{next_fifth[2:4]},{next_sixth[2:4]}  蓝球：{next_basketball[2:4]}')


if __name__ == '__main__':
    a = TwoColorBall()
    a.main()